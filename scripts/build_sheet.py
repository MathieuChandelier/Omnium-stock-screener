"""scripts/build_sheet.py — le miroir .xlsx du moteur.

Le classeur ne contient PAS des valeurs : il contient les FORMULES du moteur,
ecrites sur des variables NOMMEES. C'est le meme modele, lisible a l'oeil.

Chaque ligne calculee porte, a cote de sa formule, la valeur produite par
engine/model.py. La feuille Controles compare les deux : si une formule
diverge du moteur, Excel le dit a l'ouverture. Le classeur est donc son
propre test de non-regression.

Usage :  python3 scripts/build_sheet.py inputs/SYNTHETIQUE.json [-o sortie.xlsx]
"""
import hashlib
import json
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

from engine.model import ENGINE_VERSION, bases, projette

NAVY = "000190"
TEAL = "2BD8C2"
GREY = "64748B"
LIGHT = "EEF0FB"
VEIL = "F2FCFA"

H1 = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
LBL = Font(name="Calibri", size=10, color="0D0D1A")
LBLB = Font(name="Calibri", size=10, bold=True, color=NAVY)
SMALL = Font(name="Calibri", size=9, color=GREY)
MONO = Font(name="Calibri", size=10)
FILL_H = PatternFill("solid", fgColor=NAVY)
FILL_L = PatternFill("solid", fgColor=LIGHT)
FILL_V = PatternFill("solid", fgColor=VEIL)
THIN = Side(style="thin", color="D8DBE8")
BOX = Border(bottom=THIN)


def _titre(ws, row, texte, largeur):
    ws.cell(row=row, column=1, value=texte).font = H1
    for c in range(1, largeur + 1):
        ws.cell(row=row, column=c).fill = FILL_H
    return row + 1


def _nom(wb, ws, nom, row, col):
    """Declare une variable nommee sur une cellule. C'est ce qui rend les
    formules lisibles : =CA_2026*marge_2026 plutot que =C12*D5."""
    ref = "'%s'!$%s$%d" % (ws.title, get_column_letter(col), row)
    wb.defined_names.add(DefinedName(nom, attr_text=ref))


EPOQUE = (2000, 1, 1, 0, 0, 0)


def _normaliser(chemin):
    """Rend le .xlsx REPRODUCTIBLE a l'octet.

    openpyxl reecrit `dcterms:modified` a l'enregistrement et le ZIP horodate
    chaque membre a l'instant de l'ecriture : deux builds identiques
    produisaient deux fichiers differents. Sans cette normalisation, une
    regeneration des 59 titres afficherait 59 fichiers modifies alors que rien
    n'a bouge, et le diff - qui est justement le rapport d'impact - devient
    illisible. On repasse donc l'archive avec un horodatage fige et un ordre
    stable.
    """
    import re
    import shutil
    import zipfile

    src = zipfile.ZipFile(chemin)
    membres = [(i.filename, src.read(i.filename)) for i in sorted(src.infolist(), key=lambda x: x.filename)]
    src.close()
    tmp = chemin + ".tmp"
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as out:
        for nom, data in membres:
            if nom == "docProps/core.xml":
                data = re.sub(rb">[^<]*</dcterms:modified>", b">2000-01-01T00:00:00Z</dcterms:modified>", data)
            info = zipfile.ZipInfo(nom, date_time=EPOQUE)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = 0o600 << 16
            out.writestr(info, data)
    shutil.move(tmp, chemin)


def construire(inp, chemin):
    h = inp["hypotheses"]
    horizon = h["horizon"]
    hist_key = sorted(inp["comptes"])[-1]
    an0 = int(hist_key[:4])
    per = inp["comptes"][hist_key]
    pub = per["publie"]
    omn, aj, coin = bases(per)
    proj = projette(inp)
    bil0 = inp["bilan"][hist_key]
    empreinte = hashlib.sha256(json.dumps(inp, sort_keys=True).encode()).hexdigest()[:12]

    wb = Workbook()
    # PURETE AU NIVEAU DU FICHIER. openpyxl horodate le classeur a la
    # generation : deux builds identiques produisaient deux fichiers
    # differents, et `git status` aurait signale les 59 comme modifies apres
    # chaque regeneration, meme sans le moindre changement. On fige donc les
    # proprietes de document - le seul horodatage qui compte est l'empreinte
    # des entrees, portee en A1.
    from datetime import datetime
    fige = datetime(2000, 1, 1)
    wb.properties.created = fige
    wb.properties.modified = fige
    wb.properties.creator = "Omnium engine v%s" % ENGINE_VERSION
    wb.properties.lastModifiedBy = "Omnium engine v%s" % ENGINE_VERSION

    # ══ 1. Lisez-moi ═══════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Lisez-moi"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 96
    ws["A1"] = "%s — moteur v%s — empreinte %s" % (inp["identite"]["nom"], ENGINE_VERSION, empreinte)
    ws["A1"].font = H1
    ws["A1"].fill = FILL_H
    ws.merge_cells("A1:B1")
    lignes = [
        ("", ""),
        ("Ce que contient ce fichier",
         "Les FORMULES du moteur, pas ses resultats. Chaque cellule calculee est une formule "
         "ecrite sur des variables nommees, lisible et modifiable."),
        ("Le test integre",
         "Colonne 'moteur' = la valeur produite par engine/model.py. La feuille Controles compare "
         "formule et moteur ligne a ligne : un ecart s'affiche a l'ouverture."),
        ("Peremption",
         "Une hypothese de rang 'guidance' ne vaut QUE pour son exercice. Au-dela le moteur bascule "
         "sur le normatif - visible dans la colonne rang de la feuille Modele."),
        ("Rang du resultat financier",
         "1 = detail publie par instrument. 2 = taux appliques aux postes de bilan decomposes. "
         "3 = spread normatif sur la dette nette. La valeur est retournee AVEC son rang."),
        ("Base Omnium",
         "publie + vrais one-offs nets d'impot. L'ajuste du management ajoute en plus les "
         "recurrents-deguises : c'est le coin, et Omnium le GARDE."),
        ("Si l'empreinte ne correspond plus",
         "Les entrees ont change depuis la generation : cette feuille est perimee, regenerez-la."),
        ("", ""),
        ("Devise de reporting", inp["identite"]["reportingCurrency"]),
        ("Devise de cotation", inp["identite"]["quoteCurrency"]),
        ("Cours", "%s au %s" % (inp["identite"]["cours"], inp["identite"]["coursAsOf"])),
    ]
    r = 2
    for a, b in lignes:
        ws.cell(row=r, column=1, value=a).font = LBLB
        c = ws.cell(row=r, column=2, value=b)
        c.font = LBL
        c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    # ══ 2. Hypotheses ══════════════════════════════════════════════════
    ws = wb.create_sheet("Hypotheses")
    for col, w in zip("ABCDEF", (30, 12, 12, 12, 14, 62)):
        ws.column_dimensions[col].width = w
    r = _titre(ws, 1, "MOTEURS — toute cellule projetee est pilotee par l'un d'eux", 6)
    for i, t in enumerate(["moteur", "variable", "valeur", "rang", "expire", "justification / source"], start=1):
        c = ws.cell(row=r, column=i, value=t)
        c.font = LBLB
        c.fill = FILL_L
    r += 1

    def pose(nom_lisible, var, val, rang, expire, just):
        nonlocal r
        ws.cell(row=r, column=1, value=nom_lisible).font = LBL
        ws.cell(row=r, column=2, value=var).font = SMALL
        c = ws.cell(row=r, column=3, value=val)
        c.font = MONO
        c.number_format = "0.000"
        ws.cell(row=r, column=4, value=rang).font = SMALL
        ws.cell(row=r, column=5, value=expire if expire else "—").font = SMALL
        j = ws.cell(row=r, column=6, value=just)
        j.font = SMALL
        j.alignment = Alignment(wrap_text=True, vertical="top")
        for cc in range(1, 7):
            ws.cell(row=r, column=cc).border = BOX
        _nom(wb, ws, var, r, 3)
        r += 1

    for an in horizon:
        e = h["croissanceCA"].get(str(an), {})
        pose("Croissance du CA %d" % an, "g_%d" % an, e.get("valeur"), e.get("rang"), e.get("expire"), e.get("moteur", ""))
    for an in horizon:
        e = h["margeEbitda"].get(str(an), {})
        pose("Marge d'EBITDA %d" % an, "marge_%d" % an, e.get("valeur"), e.get("rang"), e.get("expire"), e.get("moteur", ""))
    for an in horizon:
        e = h["tauxImpot"].get(str(an)) or h["tauxImpot"]["defaut"]
        applique = e if (str(an) in h["tauxImpot"] and (e.get("expire") is None or an <= e["expire"])) else h["tauxImpot"]["defaut"]
        pose("Taux d'impot %d" % an, "tauxIS_%d" % an, applique["valeur"], applique.get("rang"),
             applique.get("expire"), applique.get("moteur", ""))
    pose("Capex / CA", "capexRatio", h["capexSurCA"]["valeur"], "normatif", None, h["capexSurCA"]["moteur"])
    pose("Duree d'amortissement", "dureeAmort", h["dureeAmortissement"]["valeur"], "normatif", None, h["dureeAmortissement"]["moteur"])
    rf = h["resultatFinancier"]
    if "tauxDetteBrute" in rf:
        pose("Taux sur dette brute", "tauxDette", rf["tauxDetteBrute"], "rang %d" % rf["rang"], None, rf["moteur"])
    if "tauxPlacementCash" in rf:
        pose("Taux sur tresorerie", "tauxCash", rf["tauxPlacementCash"], "rang %d" % rf["rang"], None, rf["moteur"])
    if "tauxLease" in rf:
        pose("Taux sur dette locative", "tauxLease", rf["tauxLease"], "rang %d" % rf["rang"], None, rf["moteur"])
    if "spreadNormatif" in rf:
        pose("Spread normatif sur dette nette", "spreadNormatif", rf["spreadNormatif"],
             "rang %d" % rf["rang"], None, rf["moteur"])
    if "valeur" in rf:
        pose("Resultat financier publie", "finPublie", rf["valeur"], "rang %d" % rf["rang"], None, rf["moteur"])
    pose("Rachat d'actions / an", "rachatPct", h["actions"]["rachatAnnuelPct"], h["actions"]["rang"],
         h["actions"].get("expire"), h["actions"]["moteur"])
    pose("Taux de distribution", "tauxDistrib", h["distribution"]["tauxDistribution"], "normatif", None, h["distribution"]["moteur"])
    pose("Croissance terminale", "gTerminal", h["regleTerminale"]["croissanceTerminale"], "normatif", None, h["regleTerminale"]["moteur"])
    pose("Marge terminale", "margeTerminale", h["regleTerminale"]["margeTerminale"], "normatif", None, h["regleTerminale"]["moteur"])
    pose("Minoritaires / resultat", "partMinos", pub["minoritaires"] / pub["net"], "normatif", None,
         "Part des minoritaires observee au dernier exercice publie, reconduite")

    # ══ 3. Historique ══════════════════════════════════════════════════
    ws = wb.create_sheet("Historique")
    for col, w in zip("ABCDEFG", (34, 13, 13, 13, 13, 13, 46)):
        ws.column_dimensions[col].width = w
    periodes = sorted(inp["comptes"])
    r = _titre(ws, 1, "ECHELLE PUBLIEE — un modele unique, les barreaux non publies restent vides", 2 + len(periodes))
    ECH = ["ca", "margeBrute", "ebitda", "da", "ebit", "finNet", "avantImpot", "impot", "minoritaires", "net"]
    ws.cell(row=r, column=1, value="barreau").font = LBLB
    ws.cell(row=r, column=1).fill = FILL_L
    for i, pk in enumerate(periodes):
        c = ws.cell(row=r, column=2 + i, value=pk)
        c.font = LBLB
        c.fill = FILL_L
    r += 1
    col_courante = 2 + periodes.index(hist_key)
    ligne_pub = {}
    for k in ECH:
        ws.cell(row=r, column=1, value=k).font = LBL
        for i, pk in enumerate(periodes):
            v = inp["comptes"][pk]["publie"].get(k)
            if isinstance(v, (int, float)):
                c = ws.cell(row=r, column=2 + i, value=v)
                c.font = MONO
                c.number_format = "#,##0.0"
        if isinstance(pub.get(k), (int, float)):
            _nom(wb, ws, "pub_%s" % k, r, col_courante)
            ligne_pub[k] = r
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="Taux d'impot effectif observe").font = LBLB
    for i, pk in enumerate(periodes):
        p2 = inp["comptes"][pk]["publie"]
        if p2.get("avantImpot"):
            c = ws.cell(row=r, column=2 + i, value=-p2["impot"] / p2["avantImpot"])
            c.font = MONO
            c.number_format = "0.0%"
    r += 1

    r += 1
    ws.cell(row=r, column=1, value="IDENTITES DE L'ECHELLE").font = LBLB
    r += 1
    for termes, cible in (("ebitda", "da", "ebit"), ("ebit", "finNet", "avantImpot")), :
        pass
    for termes, cible in [(("ebitda", "da"), "ebit"), (("ebit", "finNet"), "avantImpot"),
                          (("avantImpot", "impot", "minoritaires"), "net")]:
        if cible in ligne_pub and all(t in ligne_pub for t in termes):
            ws.cell(row=r, column=1, value=" + ".join(termes) + " = " + cible).font = SMALL
            f = "=" + "+".join("pub_%s" % t for t in termes) + "-pub_%s" % cible
            c = ws.cell(row=r, column=col_courante, value=f)
            c.font = MONO
            c.number_format = "0.00"
            ws.cell(row=r, column=col_courante + 1,
                    value='=IF(ABS(%s%d)<=MAX(1,ABS(pub_%s)*0.005),"OK","ECART")'
                    % (get_column_letter(col_courante), r, cible)).font = LBLB
            r += 1

    r += 1
    r = _titre(ws, r, "RETRAITEMENTS — la table de l'emetteur, ligne a ligne", 7)
    for i, t in enumerate(["ligne", "ebitda", "ebit", "net", "classe", "origine", "libelle"], start=1):
        c = ws.cell(row=r, column=i, value=t)
        c.font = LBLB
        c.fill = FILL_L
    r += 1
    debut_ret = r
    for rt in per.get("retraitements", []):
        ws.cell(row=r, column=1, value=rt["id"]).font = LBL
        for j, k in enumerate(("ebitda", "ebit", "net"), start=2):
            c = ws.cell(row=r, column=j, value=rt.get(k, 0.0))
            c.font = MONO
            c.number_format = "#,##0.0"
        ws.cell(row=r, column=5, value=rt.get("classe")).font = SMALL
        ws.cell(row=r, column=6, value=rt.get("origine", "emetteur")).font = SMALL
        ws.cell(row=r, column=7, value=rt.get("libelle")).font = SMALL
        r += 1
    fin_ret = r - 1
    c = ws.cell(row=r, column=1, value="impot sur retraitements")
    c.font = LBL
    ws.cell(row=r, column=4, value=per.get("impotSurRetraitements", 0.0)).font = MONO
    _nom(wb, ws, "impotRet", r, 4)
    ligne_impot = r
    r += 2

    r = _titre(ws, r, "LES TROIS BASES — calculees, jamais saisies", 7)
    ws.cell(row=r, column=1, value="base").font = LBLB
    for j, k in enumerate(("ebitda", "ebit", "net"), start=2):
        ws.cell(row=r, column=j, value=k).font = LBLB
    ws.cell(row=r, column=5, value="moteur (net)").font = LBLB
    for cc in range(1, 6):
        ws.cell(row=r, column=cc).fill = FILL_L
    r += 1
    for nom, cls, val in (("Omnium (publie + vrais one-offs)", "oneOff", omn),
                          ("Ajuste du management (+ recurrents)", None, aj),
                          ("Coin = recurrents-deguises", "recurrent", coin)):
        ws.cell(row=r, column=1, value=nom).font = LBL
        for j, k in enumerate(("ebitda", "ebit", "net"), start=2):
            col = get_column_letter(j)
            if cls:
                f = ('=SUMIF($E$%d:$E$%d,"%s",%s$%d:%s$%d)' % (debut_ret, fin_ret, cls, col, debut_ret, col, fin_ret))
                if k == "net":
                    f += '+impotRet*SUMIF($E$%d:$E$%d,"%s",$D$%d:$D$%d)/SUM($D$%d:$D$%d)' % (
                        debut_ret, fin_ret, cls, debut_ret, fin_ret, debut_ret, fin_ret)
                f = ("=pub_%s+" % k) + f[1:] if cls == "oneOff" else f
            else:
                f = "=pub_%s+SUM(%s$%d:%s$%d)" % (k, col, debut_ret, col, fin_ret) + ("+impotRet" if k == "net" else "")
            c = ws.cell(row=r, column=j, value=f)
            c.font = MONO
            c.number_format = "#,##0.0"
        ws.cell(row=r, column=5, value=round(val.get("net", 0.0), 2)).font = SMALL
        r += 1
    ws.cell(row=r, column=1, value="ajuste PUBLIE par la societe (cible de bouclage)").font = LBLB
    for j, k in enumerate(("ebitda", "ebit", "net"), start=2):
        v = (per.get("ajustePublie") or {}).get(k)
        if v is not None:
            c = ws.cell(row=r, column=j, value=v)
            c.font = MONO
            c.number_format = "#,##0.0"
            _nom(wb, ws, "ajPub_%s" % k, r, j)
    r += 1
    ws.cell(row=r, column=1, value="BOUCLAGE").font = LBLB
    for j, k in enumerate(("ebitda", "ebit", "net"), start=2):
        if (per.get("ajustePublie") or {}).get(k) is not None:
            col = get_column_letter(j)
            ws.cell(row=r, column=j,
                    value='=IF(ABS(%s%d-ajPub_%s)<=MAX(1,ABS(ajPub_%s)*0.005),"OK","NE BOUCLE PAS")' % (col, r - 2, k, k)).font = LBLB

    lignes = _construire_modele(wb, inp, proj, pub, bil0, an0, horizon)
    _construire_controles(wb, proj, horizon, lignes)
    wb.save(chemin)
    _normaliser(chemin)
    return empreinte, len(proj)


def _construire_modele(wb, inp, proj, pub, bil0, an0, horizon):
    """La projection. Une colonne par annee, une formule par cellule."""
    h = inp["hypotheses"]
    ws = wb.create_sheet("Modele")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 13
    for i in range(len(horizon)):
        ws.column_dimensions[get_column_letter(3 + i)].width = 13
    ws.column_dimensions[get_column_letter(3 + len(horizon))].width = 3
    for i in range(len(horizon)):
        ws.column_dimensions[get_column_letter(4 + len(horizon) + i)].width = 11

    r = _titre(ws, 1, "MODELE — chaque cellule est une formule sur des variables nommees", 3 + len(horizon))
    ws.cell(row=r, column=1, value="ligne").font = LBLB
    ws.cell(row=r, column=2, value=an0).font = LBLB
    ws.cell(row=r, column=2).fill = FILL_L
    ws.cell(row=r, column=1).fill = FILL_L
    for i, an in enumerate(horizon):
        c = ws.cell(row=r, column=3 + i, value=an)
        c.font = LBLB
        c.fill = FILL_V
    ws.cell(row=r, column=4 + len(horizon), value="← moteur (test)").font = SMALL
    for i, an in enumerate(horizon):
        ws.cell(row=r, column=4 + len(horizon) + i, value=an).font = SMALL
    r += 1
    entete = r - 1

    lignes = {}
    formules = []

    def ligne(label, base_val, formule, cle_moteur, fmt="#,##0.0", nom=None, gras=False, symbole=None):
        nonlocal r
        ws.cell(row=r, column=1, value=label).font = LBLB if gras else LBL
        if base_val is not None:
            c = ws.cell(row=r, column=2, value=base_val)
            c.font = MONO
            c.number_format = fmt
            if nom:
                _nom(wb, ws, "%s_%d" % (nom, an0), r, 2)
        for i, an in enumerate(horizon):
            col = get_column_letter(3 + i)
            prev = get_column_letter(2 + i)
            c = ws.cell(row=r, column=3 + i, value=formule(an, col, prev, r))
            c.font = MONO
            c.number_format = fmt
            c.fill = FILL_V
            if nom:
                _nom(wb, ws, "%s_%d" % (nom, an), r, 3 + i)
        if cle_moteur:
            for i, p in enumerate(proj):
                c = ws.cell(row=r, column=4 + len(horizon) + i, value=round(p[cle_moteur], 4))
                c.font = SMALL
                c.number_format = fmt
        lignes[label] = r
        if symbole:
            an1 = horizon[0]
            formules.append((label, symbole, ws.cell(row=r, column=3).value))
        r += 1
        return r - 1

    # — compte de resultat —
    l_ca = ligne("Chiffre d'affaires", pub["ca"],
                 lambda an, col, prev, rr: "=%s%d*(1+g_%d)" % (prev, rr, an), "ca", nom="CA", gras=True, symbole="CA(n) = CA(n-1) x (1 + g_n)")
    ligne("Croissance", None, lambda an, col, prev, rr: "=g_%d" % an, "croissance", fmt="0.0%")
    l_marge = ligne("Marge d'EBITDA", None, lambda an, col, prev, rr: "=marge_%d" % an, "margeEbitda", fmt="0.0%")
    l_ebitda = ligne("EBITDA", pub["ebitda"],
                     lambda an, col, prev, rr: "=%s%d*%s%d" % (col, l_ca, col, l_marge), "ebitda", gras=True, symbole="EBITDA(n) = CA(n) x marge_n")

    da0 = -pub["da"] - sum((v.get("echeancier") or {}).get(str(an0), 0.0)
                           for k, v in (inp.get("ppa") or {}).items() if not k.startswith("_"))
    l_dahp = ligne("  dotations hors PPA", da0,
                   lambda an, col, prev, rr: "=%s%d+%s%d*capexRatio/dureeAmort" % (prev, rr, prev, l_ca), "daHorsPpa", symbole="dotHorsPPA(n) = dotHorsPPA(n-1) + CA(n-1) x capexRatio / dureeAmort")
    l_dappa = ligne("  amortissement du PPA", sum((v.get("echeancier") or {}).get(str(an0), 0.0)
                                                  for k, v in (inp.get("ppa") or {}).items() if not k.startswith("_")),
                    lambda an, col, prev, rr: "=%s" % sum((v.get("echeancier") or {}).get(str(an), 0.0)
                                                          for k, v in (inp.get("ppa") or {}).items() if not k.startswith("_")),
                    "daPpa")
    l_da = ligne("Dotations", pub["da"],
                 lambda an, col, prev, rr: "=-(%s%d+%s%d)" % (col, l_dahp, col, l_dappa), "da", symbole="dotations(n) = -( dotHorsPPA(n) + amortPPA(n) )")
    l_ebit = ligne("EBIT", pub["ebit"],
                   lambda an, col, prev, rr: "=%s%d+%s%d" % (col, l_ebitda, col, l_da), "ebit", gras=True, symbole="EBIT(n) = EBITDA(n) + dotations(n)")

    # — bilan et resultat financier (cascade de rang) —
    l_dette = ligne("Dette brute", bil0["debtGross"], lambda an, col, prev, rr: "=%s%d" % (prev, rr), None, nom="dette")
    l_cash = ligne("Tresorerie", bil0["cash"], lambda an, col, prev, rr: "=%s%d" % (prev, rr), None, nom="cash")
    l_lease = ligne("Dette locative", bil0["leaseDebt"], lambda an, col, prev, rr: "=%s%d" % (prev, rr), None)
    rang_fin = h["resultatFinancier"].get("rang", 3)
    if rang_fin == 1:
        f_fin = lambda an, col, prev, rr: "=finPublie"
    elif rang_fin == 2:
        f_fin = lambda an, col, prev, rr: ("=-(%s%d*tauxDette)+%s%d*tauxCash-(%s%d*tauxLease)"
                                           % (prev, l_dette, prev, l_cash, prev, l_lease))
    else:
        f_fin = lambda an, col, prev, rr: ("=-(%s%d-%s%d+%s%d)*spreadNormatif"
                                           % (prev, l_dette, prev, l_cash, prev, l_lease))
    SYM_FIN = {1: "resultatFinancier(n) = finPublie   [rang 1 : detail publie par instrument]",
               2: "resultatFinancier(n) = -detteBrute(n-1) x tauxDette + tresorerie(n-1) x tauxCash - detteLocative(n-1) x tauxLease   [rang 2]",
               3: "resultatFinancier(n) = -( detteBrute(n-1) - tresorerie(n-1) + detteLocative(n-1) ) x spreadNormatif   [rang 3]"}
    l_fin = ligne("Resultat financier (rang %d)" % rang_fin, pub["finNet"], f_fin, "finNet",
                  symbole=SYM_FIN[rang_fin])

    l_avant = ligne("Resultat avant impot", pub["avantImpot"],
                    lambda an, col, prev, rr: "=%s%d+%s%d" % (col, l_ebit, col, l_fin), "avantImpot", symbole="avantImpot(n) = EBIT(n) + resultatFinancier(n)")
    l_taux = ligne("Taux d'impot applique", None, lambda an, col, prev, rr: "=tauxIS_%d" % an, "tauxImpot", fmt="0.0%")
    # OVERRIDE DECLARE (regle 7). Le moteur calcule ; quand le jugement
    # diverge, la valeur posee dans inputs/ remplace la formule - et elle
    # apparait TELLE QUELLE dans la feuille, avec sa justification en
    # commentaire de cellule. Une exception declaree, jamais silencieuse.
    ovr = {int(a): v for a, v in (inp.get("overrides") or {}).items() if a.isdigit()}

    def f_impot(an, col, prev, rr):
        o = (ovr.get(an) or {}).get("impot")
        return o["valeur"] if o else "=-%s%d*%s%d" % (col, l_avant, col, l_taux)

    l_imp = ligne("Impot", pub["impot"], f_impot, "impot",
                  symbole="impot(n) = -avantImpot(n) x tauxIS_n   [ou OVERRIDE declare]")
    for i, an in enumerate(horizon):
        o = (ovr.get(an) or {}).get("impot")
        if o:
            c = ws.cell(row=l_imp, column=3 + i)
            c.fill = PatternFill("solid", fgColor="FFF3CD")
            c.comment = Comment("OVERRIDE DECLARE\n\n" + o["justification"], "Omnium engine")
    l_min = ligne("Minoritaires", pub["minoritaires"],
                  lambda an, col, prev, rr: "=partMinos*(%s%d+%s%d)" % (col, l_avant, col, l_imp), "minoritaires", symbole="minoritaires(n) = partMinos x ( avantImpot(n) + impot(n) )")
    l_net = ligne("Resultat net", pub["net"],
                  lambda an, col, prev, rr: "=%s%d+%s%d+%s%d" % (col, l_avant, col, l_imp, col, l_min), "net", gras=True, symbole="net(n) = avantImpot(n) + impot(n) + minoritaires(n)")

    l_act = ligne("Actions (millions)", inp["actions"][str(an0)],
                  lambda an, col, prev, rr: "=%s%d*(1-rachatPct)" % (prev, rr), "actions", fmt="#,##0.00", symbole="actions(n) = actions(n-1) x (1 - rachatPct)")
    l_eps = ligne("BPA", pub["net"] / inp["actions"][str(an0)],
                  lambda an, col, prev, rr: "=%s%d/%s%d" % (col, l_net, col, l_act), "eps", fmt="#,##0.00", gras=True, symbole="BPA(n) = net(n) / actions(n)")
    l_div = ligne("Dividende verse", None,
                  lambda an, col, prev, rr: "=%s%d*tauxDistrib" % (col, l_net), "dividende", symbole="dividende(n) = net(n) x tauxDistrib")

    # la tresorerie roule apres coup : elle depend du resultat de l'annee
    for i, an in enumerate(horizon):
        col = get_column_letter(3 + i)
        prev = get_column_letter(2 + i)
        # Les dotations sont une charge NON CASH dans leur totalite, jambe PPA
        # comprise : on reprend la ligne Dotations (negative), pas la seule
        # jambe hors PPA. C'est l'ecart que le test integre a leve.
        ws.cell(row=l_cash, column=3 + i,
                value="=%s%d+%s%d-%s%d-%s%d*capexRatio-%s%d" % (prev, l_cash, col, l_net, col, l_div, prev, l_ca, col, l_da))

    r += 1
    ws.cell(row=r, column=1, value="RANG DU MOTEUR — d'ou vient chaque hypothese").font = LBLB
    r += 1
    for label, cle in (("croissance", "rangCroissance"), ("marge", "rangMarge"),
                       ("taux d'impot", "rangTaux"), ("resultat financier", "rangFinNet")):
        ws.cell(row=r, column=1, value="  " + label).font = SMALL
        for i, p in enumerate(proj):
            c = ws.cell(row=r, column=3 + i, value=str(p[cle]))
            c.font = SMALL
            if str(p[cle]) == "guidance":
                c.fill = PatternFill("solid", fgColor="E6FAF8")
        r += 1
    ws.freeze_panes = ws.cell(row=entete + 1, column=3)
    _feuille_formules(wb, formules, horizon)
    return lignes


def _feuille_formules(wb, formules, horizon):
    """Le modele en clair, sans cliquer cellule par cellule.

    Une spreadsheet ou il faut ouvrir chaque cellule pour savoir ce qu'elle
    calcule n'est pas un document qu'on lit. Cette feuille donne, pour chaque
    ligne : l'ecriture symbolique du calcul, puis la formule Excel telle
    qu'elle est posee sur la premiere annee projetee.
    """
    ws = wb.create_sheet("Formules", 3)
    for col, w in zip("ABC", (26, 84, 46)):
        ws.column_dimensions[col].width = w
    r = _titre(ws, 1, "FORMULES DU MODELE — l'ecriture du calcul, ligne par ligne", 3)
    for i, t in enumerate(["ligne", "calcul", "formule posee sur %d" % horizon[0]], start=1):
        c = ws.cell(row=r, column=i, value=t)
        c.font = LBLB
        c.fill = FILL_L
    r += 1
    for label, sym, excel in formules:
        ws.cell(row=r, column=1, value=label.strip()).font = LBLB
        c = ws.cell(row=r, column=2, value=sym)
        c.font = MONO
        c.alignment = Alignment(wrap_text=True, vertical="top")
        e = ws.cell(row=r, column=3, value="'" + str(excel) if isinstance(excel, str) else excel)
        e.font = SMALL
        for cc in range(1, 4):
            ws.cell(row=r, column=cc).border = BOX
        ws.row_dimensions[r].height = 28
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="OU LIRE LE RESTE").font = LBLB
    r += 1
    for a, b in [
        ("Hypotheses", "la valeur de chaque moteur, son rang de provenance et sa peremption"),
        ("Historique", "l'echelle publiee de chaque exercice, ses identites, et les trois bases"),
        ("Modele", "le deroule annee par annee ; la colonne de droite porte la valeur du moteur Python"),
        ("Controles", "formule contre moteur, cellule a cellule"),
        ("engine/model.py", "la SOURCE des formules ci-dessus — le classeur n'en est que le miroir"),
    ]:
        ws.cell(row=r, column=1, value=a).font = LBL
        ws.cell(row=r, column=2, value=b).font = SMALL
        r += 1


def _construire_controles(wb, proj, horizon, lignes):
    """Compare, cellule a cellule, la FORMULE de la feuille et la valeur du
    MOTEUR. Un ecart signifie que la traduction en formule a derive du code :
    le classeur se denonce lui-meme a l'ouverture."""
    ws = wb.create_sheet("Controles")
    ws.column_dimensions["A"].width = 40
    for i in range(len(horizon) + 1):
        ws.column_dimensions[get_column_letter(2 + i)].width = 14
    n = len(horizon)
    r = _titre(ws, 1, "CONTROLES — la feuille est son propre test de non-regression", 1 + n)
    ws.cell(row=r, column=1, value="formule vs moteur").font = LBLB
    for i, an in enumerate(horizon):
        c = ws.cell(row=r, column=2 + i, value=an)
        c.font = LBLB
        c.fill = FILL_L
    r += 1
    for label in ("Chiffre d'affaires", "EBITDA", "EBIT", "Resultat financier",
                  "Resultat avant impot", "Impot", "Resultat net", "BPA"):
        lr = lignes.get(label)
        if not lr:
            continue
        ws.cell(row=r, column=1, value=label).font = LBL
        for i in range(n):
            col = get_column_letter(3 + i)
            colm = get_column_letter(4 + n + i)
            ws.cell(row=r, column=2 + i,
                    value='=IF(ABS(Modele!%s%d-Modele!%s%d)<0.01,"OK","ECART "&TEXT(Modele!%s%d-Modele!%s%d,"0.00"))'
                    % (col, lr, colm, lr, col, lr, colm, lr)).font = MONO
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="GARDE-FOUS").font = LBLB
    r += 1
    gardes = [
        ("Aucune cellule orpheline (regle 5)",
         '=IF(COUNTBLANK(Hypotheses!$C$3:$C$%d)=0,"OK","MOTEUR MANQUANT")' % (2 + 5 * len(horizon))),
        ("Devise de reporting = devise de cotation (prop. 6)",
         '=IF(EXACT(\'Lisez-moi\'!B10,\'Lisez-moi\'!B11),"OK","P/E INTERDIT SANS TAUX DATE")'),
        ("Bouclage de la table de l'emetteur",
         '=Historique!B%d' % 0),
    ]
    for lbl, f in gardes[:2]:
        ws.cell(row=r, column=1, value=lbl).font = LBL
        ws.cell(row=r, column=2, value=f).font = MONO
        r += 1
    ws.cell(row=r + 1, column=1,
            value="Colonne 'moteur' de la feuille Modele = valeur produite par engine/model.py. "
                  "Toute divergence est une erreur de traduction, pas une tolerance.").font = SMALL


def main():
    """--all genere une feuille par jeu d'entrees ; sinon un seul titre.

    Le classeur est une SORTIE de build, jamais une source : une correction se
    porte dans inputs/ (un moteur) ou dans engine/ (une formule), jamais dans
    la feuille - une modification faite dans Excel serait perdue au build
    suivant. C'est ce qui garantit qu'il n'existe qu'une seule version du
    modele.
    """
    args = [a for a in sys.argv[1:] if not a.startswith("-")]
    dossier = "artifacts"
    if "-o" in sys.argv:
        dossier = None

    if "--all" in sys.argv:
        sources = sorted(
            os.path.join("inputs", f) for f in os.listdir("inputs") if f.endswith(".json"))
        if not sources:
            print("aucun jeu d'entrees dans inputs/")
            return 2
        os.makedirs(dossier, exist_ok=True)
        for src in sources:
            nom = os.path.basename(src).replace(".json", "")
            out = os.path.join(dossier, nom + ".xlsx")
            inp = json.load(open(src, encoding="utf-8"))
            emp, n = construire(inp, out)
            print("  %-22s %d annees  empreinte %s" % (nom + ".xlsx", n, emp))
        print("%d feuille(s) — moteur v%s" % (len(sources), ENGINE_VERSION))
        return 0

    if not args:
        print(__doc__)
        return 2
    src = args[0]
    out = ("%s/%s.xlsx" % (dossier, os.path.basename(src).replace(".json", ""))
           if dossier else sys.argv[sys.argv.index("-o") + 1])
    os.makedirs(os.path.dirname(out) or ".", exist_ok=True)
    inp = json.load(open(src, encoding="utf-8"))
    emp, n = construire(inp, out)
    print("%s ecrit — %d annees projetees, empreinte %s, moteur v%s" % (out, n, emp, ENGINE_VERSION))
    return 0


if __name__ == "__main__":
    sys.exit(main())
